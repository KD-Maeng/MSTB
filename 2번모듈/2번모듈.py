import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.page import PageMargins

# 엑셀 파일 읽기
input_file = 'input.xlsx'
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# 새로운 워크북 생성
report_wb = Workbook()
report_ws = report_wb.active

# A4 용지 설정
report_ws.page_setup.paperSize = report_ws.PAPERSIZE_A4
report_ws.page_setup.orientation = report_ws.ORIENTATION_PORTRAIT
report_ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

# 폰트 및 스타일 설정
header_font = Font(bold=True, size=14)
content_font = Font(size=12)
title_font = Font(bold=True, size=16)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 데이터 읽기
headers = [cell.value for cell in ws[1]]
data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]

# 보고서 작성
for row_idx, row_data in enumerate(data, start=1):
    start_row = (row_idx - 1) * 42 + 1
    
    # 제목
    report_ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    report_ws.cell(row=start_row, column=1, value="2025 NEWSONG J 그룹배치").font = title_font
    report_ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    report_ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=6)
    report_ws.cell(row=start_row + 1, column=1, value="오직주님, 칠년을 하루같이").font = title_font
    report_ws.cell(row=start_row + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    report_ws.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=6)
    report_ws.cell(row=start_row + 2, column=1, value="그룹원 소견서").font = title_font
    report_ws.cell(row=start_row + 2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # 기본 정보
    report_ws.cell(row=start_row + 4, column=1, value="대상자").font = header_font
    report_ws.cell(row=start_row + 4, column=3, value="이름").font = header_font
    report_ws.cell(row=start_row + 4, column=4, value=row_data[2]).font = content_font
    report_ws.cell(row=start_row + 4, column=5, value="성별").font = header_font
    report_ws.cell(row=start_row + 4, column=6, value=row_data[5]).font = content_font
    
    report_ws.cell(row=start_row + 5, column=1, value="교인구분").font = header_font
    report_ws.cell(row=start_row + 5, column=3, value="기수").font = header_font
    report_ws.cell(row=start_row + 5, column=4, value=row_data[4]).font = content_font
    report_ws.cell(row=start_row + 5, column=5, value="생년월일").font = header_font
    report_ws.cell(row=start_row + 5, column=6, value=row_data[6]).font = content_font
    
    report_ws.cell(row=start_row + 6, column=1, value="학교/직장").font = header_font
    report_ws.cell(row=start_row + 6, column=3, value="전공").font = header_font
    
    report_ws.cell(row=start_row + 7, column=1, value="현재상황").font = header_font
    report_ws.cell(row=start_row + 7, column=3, value="예정사항").font = header_font
    
    report_ws.cell(row=start_row + 8, column=1, value="전화번호").font = header_font
    report_ws.cell(row=start_row + 8, column=4, value=row_data[3]).font = content_font
    
    # 출석 정보
    report_ws.cell(row=start_row + 10, column=1, value="출석정보").font = header_font
    report_ws.cell(row=start_row + 10, column=2, value="뉴송예배").font = header_font
    report_ws.cell(row=start_row + 10, column=3, value=row_data[13]).font = content_font
    report_ws.cell(row=start_row + 10, column=4, value="주일낮").font = header_font
    report_ws.cell(row=start_row + 10, column=5, value=row_data[14]).font = content_font
    report_ws.cell(row=start_row + 10, column=6, value="주일저녁").font = header_font
    report_ws.cell(row=start_row + 10, column=7, value=row_data[15]).font = content_font
    
    # 소견서
    report_ws.cell(row=start_row + 12, column=1, value="소견서").font = header_font
    report_ws.merge_cells(start_row=start_row + 12, start_column=2, end_row=start_row + 20, end_column=6)
    report_ws.cell(row=start_row + 12, column=2, value=row_data[7]).font = content_font
    report_ws.cell(row=start_row + 12, column=2).alignment = Alignment(horizontal='left', vertical='top')
    
    # 동반배치 사유
    report_ws.cell(row=start_row + 22, column=1, value="동반배치 사유").font = header_font
    report_ws.merge_cells(start_row=start_row + 22, start_column=2, end_row=start_row + 24, end_column=6)
    report_ws.cell(row=start_row + 22, column=2, value=row_data[17]).font = content_font
    report_ws.cell(row=start_row + 22, column=2).alignment = Alignment(horizontal='left', vertical='top')
    
    # 소견서 작성자
    report_ws.cell(row=start_row + 26, column=1, value="소견서 작성자").font = header_font
    report_ws.cell(row=start_row + 26, column=2, value="이름").font = header_font
    report_ws.cell(row=start_row + 26, column=3, value=row_data[18]).font = content_font
    report_ws.cell(row=start_row + 26, column=4, value="연락처").font = header_font
    report_ws.cell(row=start_row + 26, column=5, value=row_data[19]).font = content_font

# 열 너비 자동 조정
for col in report_ws.columns:
    max_length = 0
    column = None
    for cell in col:
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            if column is None:
                column = cell.column_letter
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
    if column:
        adjusted_width = (max_length + 2)
        report_ws.column_dimensions[column].width = adjusted_width

# 보고서 저장
output_file = 'report.xlsx'
report_wb.save(output_file)

print(f"보고서가 {output_file}로 저장되었습니다.")